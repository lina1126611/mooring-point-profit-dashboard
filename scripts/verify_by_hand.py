"""엔진 결과를 원본 엑셀 수기 집계와 대조한다.

    python scripts/verify_by_hand.py ["프로젝트명"]

검증 경로를 둘로 갈라 놓는 것이 핵심이다.
    (A) 엔진 : 엑셀 → ingest → normalize → SQLite → profit/finance
    (B) 수기 : 엑셀 → pandas 필터·합계 (= 사람이 엑셀에서 하는 것과 같은 연산)
둘이 일치하면 파싱·정규화·적재·집계·배부 중간에 금액이 새지 않았다는 뜻이다.
(분류 규칙 자체는 양쪽이 공유한다 — 검증 대상은 규칙이 아니라 계산 경로다.)

마지막에 reports/ 에 검산용 엑셀을 만든다. 그 파일의 금액 칸은 전부 SUM/
SUMIFS 수식이라 원본 행을 지우거나 고치면 숫자가 따라 움직인다. 대표가
"이 숫자 어디서 나왔냐"고 물으면 그 파일을 열어 보여 주면 된다.
"""

from __future__ import annotations

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

import pandas as pd  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from load_sample_data import FIXED_COSTS, LOANS  # noqa: E402
from src import db as db_module  # noqa: E402
from src import profit  # noqa: E402
from src.classify import classify_account, classify_behavior  # noqa: E402
from src.rules import FIXED, INTEREST_ACCOUNT, VARIABLE  # noqa: E402

SAMPLE_DIR = PROJECT_ROOT / "data" / "sample"
REPORT_DIR = PROJECT_ROOT / "reports"
DB_PATH = PROJECT_ROOT / "db" / "mooring.db"

SALES_XLSX = SAMPLE_DIR / "매출_세금계산서_2026.xlsx"
PURCHASE_XLSX = SAMPLE_DIR / "매입_세금계산서_2026.xlsx"
EXPENSE_XLSX = SAMPLE_DIR / "경비지출대장_2026.xlsx"
MANDAY_XLSX = SAMPLE_DIR / "설계맨데이_투입내역.xlsx"


def won(n) -> str:
    return f"{int(round(n or 0)):,}"


def pct(x) -> str:
    return f"{x * 100:.1f}%"


def rule(title: str) -> None:
    print(f"\n{'=' * 80}\n{title}\n{'=' * 80}")


def parse_won(value) -> int:
    """'1,234,000' / 1234000 → 1234000. 엑셀에서 사람이 읽는 그대로."""
    if pd.isna(value):
        return 0
    return int(str(value).replace(",", "").replace("원", "").strip())


# ===============================================================
# (B) 수기 집계 — 원본 엑셀만 읽는다
# ===============================================================


def load_raw() -> dict[str, pd.DataFrame]:
    """원본 엑셀 4종을 읽고, 매입·경비에 계정/원가행태를 손으로 태그한 것과
    같은 열을 붙인다."""
    sales = pd.read_excel(SALES_XLSX)

    purchase = pd.read_excel(PURCHASE_XLSX)
    purchase["계정"] = [
        classify_account(v, d) for v, d in zip(purchase["거래처명"], purchase["품목"])
    ]
    purchase["원가행태"] = [classify_behavior(a) for a in purchase["계정"]]

    expense = pd.read_excel(EXPENSE_XLSX)
    expense["금액"] = expense["금액"].map(parse_won)
    expense["계정"] = [
        classify_account(v, d, a)
        for v, d, a in zip(expense["지급처"], expense["적요"], expense["계정"])
    ]
    expense["원가행태"] = [classify_behavior(a) for a in expense["계정"]]

    manday = pd.read_excel(MANDAY_XLSX)
    manday["인건비"] = manday["투입인원"] * manday["투입일수"] * manday["일단가"]

    return {"매출": sales, "매입": purchase, "경비": expense, "맨데이": manday}


def analysis_months_from_raw(raw: dict[str, pd.DataFrame]) -> int:
    """거래가 존재하는 개월 수 — 엑셀의 날짜 열 3개를 모두 훑는다."""
    stamps = pd.concat(
        [
            pd.to_datetime(raw["매출"]["발행일"]),
            pd.to_datetime(raw["매입"]["작성일자"]),
            pd.to_datetime(raw["경비"]["지출일"]),
        ]
    )
    lo, hi = stamps.min(), stamps.max()
    return (hi.year - lo.year) * 12 + (hi.month - lo.month) + 1


def hand_calc(raw: dict[str, pd.DataFrame], project: str) -> dict:
    """엑셀 필터 + 합계로 2단 손익을 직접 계산한다."""
    sales, purchase, expense, manday = raw["매출"], raw["매입"], raw["경비"], raw["맨데이"]
    months = analysis_months_from_raw(raw)

    # --- 1단 ---
    revenue = int(sales.loc[sales["현장"] == project, "공급가액"].sum())
    var_purchase = int(
        purchase.loc[
            (purchase["프로젝트"] == project) & (purchase["원가행태"] == VARIABLE), "공급가액"
        ].sum()
    )
    var_expense = int(
        expense.loc[
            (expense["현장명"] == project) & (expense["원가행태"] == VARIABLE), "금액"
        ].sum()
    )
    variable = var_purchase + var_expense

    # --- 2단 재료 ---
    manday_cost = int(round(manday.loc[manday["현장명"] == project, "인건비"].sum()))

    not_interest = expense["계정"] != INTEREST_ACCOUNT
    direct_fixed = int(
        expense.loc[
            (expense["현장명"] == project) & (expense["원가행태"] == FIXED) & not_interest, "금액"
        ].sum()
    ) + int(
        purchase.loc[
            (purchase["프로젝트"] == project) & (purchase["원가행태"] == FIXED), "공급가액"
        ].sum()
    )

    # 고정비 풀: 마스터 월액 × 개월수 + 차입금 이자 + 원장의 공통 고정비
    master = sum(f[1] for f in FIXED_COSTS) * months
    interest = sum(int(p * r / 12) for _, p, r, _, _ in LOANS) * months
    common_fixed = int(
        expense.loc[
            expense["현장명"].isna() & (expense["원가행태"] == FIXED) & not_interest, "금액"
        ].sum()
    ) + int(
        purchase.loc[
            purchase["프로젝트"].isna() & (purchase["원가행태"] == FIXED), "공급가액"
        ].sum()
    )
    pool = master + interest + common_fixed

    total_revenue = int(sales["공급가액"].sum())
    allocated = pool * revenue / total_revenue      # 매출액 비례 (잔차 보정 전)

    margin = revenue - variable
    operating = margin - allocated - direct_fixed - manday_cost

    return {
        "개월수": months,
        "매출": revenue,
        "변동비": variable,
        "변동비(매입)": var_purchase,
        "변동비(경비)": var_expense,
        "공헌이익": margin,
        "공헌이익률": margin / revenue if revenue else 0.0,
        "고정비풀": pool,
        "고정비풀_마스터": master,
        "고정비풀_이자": interest,
        "고정비풀_공통": common_fixed,
        "전사매출": total_revenue,
        "매출비중": revenue / total_revenue if total_revenue else 0.0,
        "배부고정비": allocated,
        "직접귀속고정비": direct_fixed,
        "맨데이": manday_cost,
        "진짜영업이익": operating,
        "진짜이익률": operating / revenue if revenue else 0.0,
    }


# ===============================================================
# 검산용 엑셀 (모든 금액 칸이 수식)
# ===============================================================


def write_workbook(raw: dict[str, pd.DataFrame], project: str, months: int, path: Path) -> None:
    sales = raw["매출"][["발행일", "거래처", "현장", "품목명", "공급가액"]]
    purchase = raw["매입"][["작성일자", "거래처명", "품목", "프로젝트", "계정", "원가행태", "공급가액"]]
    expense = raw["경비"][["지출일", "지급처", "적요", "현장명", "계정", "원가행태", "금액"]]
    manday = raw["맨데이"][["현장명", "직무", "투입인원", "투입일수", "일단가"]]

    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        sales.to_excel(writer, sheet_name="매출", index=False)
        purchase.to_excel(writer, sheet_name="매입", index=False)
        expense.to_excel(writer, sheet_name="경비", index=False)
        manday.to_excel(writer, sheet_name="맨데이", index=False)

        book = writer.book
        s_end, p_end, e_end, m_end = (
            len(sales) + 1, len(purchase) + 1, len(expense) + 1, len(manday) + 1
        )

        # ---------- 고정비 시트 ----------
        fx = book.create_sheet("고정비")
        fx.append(["항목", "금액", "산식"])
        fx.append(["고정비 마스터 월액", sum(f[1] for f in FIXED_COSTS),
                   " + ".join(f"{n} {a:,}" for n, a, _ in FIXED_COSTS)])
        fx.append(["분석기간(개월)", months, "거래가 존재하는 첫 달~마지막 달"])
        fx.append(["마스터 소계", "=B2*B3", "월액 × 개월수"])
        row = 5
        for name, principal, rate, _, _ in LOANS:
            fx.append([f"이자 — {name}", f"=ROUNDDOWN({principal}*{rate}/12,0)*$B$3",
                       f"원금 {principal:,} × 연 {rate * 100:.1f}% ÷ 12 × {months}개월"])
            row += 1
        interest_rows = f"B5:B{row - 1}"
        fx.append(["공통 고정비 (원장, 현장 미귀속)",
                   f'=SUMIFS(경비!G2:G{e_end},경비!D2:D{e_end},"=",경비!F2:F{e_end},"고정",'
                   f'경비!E2:E{e_end},"<>{INTEREST_ACCOUNT}")'
                   f'+SUMIFS(매입!G2:G{p_end},매입!D2:D{p_end},"=",매입!F2:F{p_end},"고정")',
                   "이자비용 행은 제외 — loans 자동계산과 이중계상되므로"])
        fx.append(["고정비 풀 (배부 대상)", f"=B4+SUM({interest_rows})+B{row}", ""])
        pool_ref = f"고정비!B{row + 1}"

        # ---------- 검산 시트 ----------
        ck = book.create_sheet("검산", 0)
        ck["A1"] = "검산 대상"
        ck["B1"] = project
        ck["A2"] = "분석기간"
        ck["B2"] = f"{months}개월"

        ck.append([])
        ck.append(["항목", "엑셀 수식 계산", "엔진 값", "차이", "산식"])

        rev = f'=SUMIF(매출!C2:C{s_end},$B$1,매출!E2:E{s_end})'
        var = (f'=SUMIFS(매입!G2:G{p_end},매입!D2:D{p_end},$B$1,매입!F2:F{p_end},"변동")'
               f'+SUMIFS(경비!G2:G{e_end},경비!D2:D{e_end},$B$1,경비!F2:F{e_end},"변동")')
        direct = (f'=SUMIFS(경비!G2:G{e_end},경비!D2:D{e_end},$B$1,경비!F2:F{e_end},"고정",'
                  f'경비!E2:E{e_end},"<>{INTEREST_ACCOUNT}")'
                  f'+SUMIFS(매입!G2:G{p_end},매입!D2:D{p_end},$B$1,매입!F2:F{p_end},"고정")')
        md = (f'=SUMPRODUCT((맨데이!$A$2:$A${m_end}=$B$1)*맨데이!$C$2:$C${m_end}'
              f'*맨데이!$D$2:$D${m_end}*맨데이!$E$2:$E${m_end})')

        lines = [
            ("매출액", rev, "매출 엑셀에서 현장 = 대상인 행의 공급가액 합계"),
            ("(−) 변동비", var, "매입·경비 중 현장 = 대상이고 원가행태 = 변동"),
            ("[1단] 공헌이익", "=B5-B6", "매출 − 변동비"),
            ("공헌이익률", "=B7/B5", "공헌이익 ÷ 매출"),
            ("(−) 배부 고정비", f"={pool_ref}*B5/SUM(매출!E2:E{s_end})",
             "고정비 풀 × (이 현장 매출 ÷ 전사 매출)"),
            ("(−) 직접귀속 고정비", direct, "이 현장에 바로 달린 고정비 (배부 대상 아님)"),
            ("(−) 맨데이 인건비", md, "Σ(투입인원 × 투입일수 × 일단가)"),
            ("[2단] 진짜 영업이익", "=B7-B9-B10-B11", "공헌이익 − 고정비 − 맨데이"),
            ("진짜이익률", "=B12/B5", "진짜 영업이익 ÷ 매출"),
            ("", "", ""),
            ("1단 − 2단 (사라진 금액)", "=B7-B12", "배부고정비 + 직접고정비 + 맨데이"),
            ("이익률 하락폭", "=B8-B13", "공헌이익률 − 진짜이익률"),
        ]
        for label, formula, note in lines:
            ck.append([label, formula, None, None, note])

        # 엔진 값과 차이 (C, D 열)
        conn = db_module.connect(DB_PATH)
        engine = next(p for p in profit.compute_all(conn) if p.project_name == project)
        conn.close()
        engine_values = {
            5: engine.revenue, 6: engine.variable_cost, 7: engine.contribution_margin,
            8: engine.contribution_margin_rate, 9: engine.allocated_fixed,
            10: engine.direct_fixed, 11: engine.manday_cost, 12: engine.operating_profit,
            13: engine.operating_profit_rate, 15: engine.gap,
            16: engine.contribution_margin_rate - engine.operating_profit_rate,
        }
        for r, value in engine_values.items():
            ck.cell(row=r, column=3, value=value)
            ck.cell(row=r, column=4, value=f"=B{r}-C{r}")

        # ---------- 서식 ----------
        head = PatternFill("solid", fgColor="DDEBF7")
        for cell in ck[4]:
            cell.font = Font(bold=True)
            cell.fill = head
        ck["A1"].font = Font(bold=True)
        for r in (7, 12):
            for c in range(1, 6):
                ck.cell(row=r, column=c).font = Font(bold=True)
        for r in range(5, 17):
            for c in (2, 3, 4):
                cell = ck.cell(row=r, column=c)
                cell.number_format = "0.0%" if r in (8, 13, 16) else "#,##0"
                cell.alignment = Alignment(horizontal="right")
        for col, width in zip("ABCDE", (26, 20, 20, 14, 52)):
            ck.column_dimensions[col].width = width
        for cell in fx[1]:
            cell.font = Font(bold=True)
            cell.fill = head
        for r in range(2, row + 2):
            fx.cell(row=r, column=2).number_format = "#,##0"
        for col, width in zip("ABC", (30, 20, 60)):
            fx.column_dimensions[col].width = width
        for sheet_name, widths in (
            ("매출", (12, 22, 30, 20, 16)),
            ("매입", (12, 20, 30, 30, 14, 10, 16)),
            ("경비", (12, 20, 24, 30, 14, 10, 14)),
            ("맨데이", (30, 14, 10, 10, 12)),
        ):
            ws = book[sheet_name]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=len(widths)).number_format = "#,##0"


# ===============================================================
def main() -> None:
    if not DB_PATH.exists():
        sys.exit("db/mooring.db 가 없다. python scripts/load_sample_data.py 를 먼저 실행할 것.")

    conn = db_module.connect(DB_PATH)
    engine_rows = profit.compute_all(conn)

    target = sys.argv[1] if len(sys.argv) > 1 else max(engine_rows, key=lambda p: p.revenue).project_name
    engine = next((p for p in engine_rows if p.project_name == target), None)
    if engine is None:
        sys.exit(f"그런 프로젝트가 없다: {target}\n"
                 + "\n".join(f"  - {p.project_name}" for p in engine_rows))

    raw = load_raw()
    hand = hand_calc(raw, target)

    # ---------------------------------------------------------
    rule(f"수기 검산 — {target}")
    print(f"  (B) 수기 : 원본 엑셀 4종을 직접 필터·합계  (분석기간 {hand['개월수']}개월)")
    print(f"  (A) 엔진 : db/mooring.db → src.profit.compute_all()")

    checks = [
        ("매출액", hand["매출"], engine.revenue, 0),
        ("변동비", hand["변동비"], engine.variable_cost, 0),
        ("공헌이익", hand["공헌이익"], engine.contribution_margin, 0),
        ("배부 고정비", hand["배부고정비"], engine.allocated_fixed, 10),
        ("직접귀속 고정비", hand["직접귀속고정비"], engine.direct_fixed, 0),
        ("맨데이 인건비", hand["맨데이"], engine.manday_cost, 0),
        ("진짜 영업이익", hand["진짜영업이익"], engine.operating_profit, 10),
    ]
    print(f"\n  {'항목':<18s}{'수기(엑셀)':>18s}{'엔진(DB)':>18s}{'차이':>12s}   판정")
    print(f"  {'-' * 74}")
    all_ok = True
    for label, h, e, tol in checks:
        diff = h - e
        ok = abs(diff) <= tol
        all_ok &= ok
        print(f"  {label:<18s}{won(h):>18s}{won(e):>18s}{won(diff):>12s}   {'✓' if ok else '✗'}")

    print(f"\n  변동비 내역: 매입 {won(hand['변동비(매입)'])} + 경비 {won(hand['변동비(경비)'])}")
    print(f"  고정비 풀  : 마스터 {won(hand['고정비풀_마스터'])} + 이자 {won(hand['고정비풀_이자'])}"
          f" + 공통 {won(hand['고정비풀_공통'])} = {won(hand['고정비풀'])}")
    print(f"  배부 몫    : {won(hand['고정비풀'])} × 매출비중 {pct(hand['매출비중'])}"
          f" = {won(hand['배부고정비'])}")
    print(f"               엔진은 {won(engine.allocated_fixed)} — 차이 "
          f"{won(engine.allocated_fixed - hand['배부고정비'])}원은 배부 잔차 보정분")
    print(f"               (내림 배부 후 남은 원 단위를 최대 가중치 현장에 몰아주어 "
          f"배부 합계 = 풀 을 보존한다)")

    # ---------------------------------------------------------
    rule("발표 장면 — 공헌이익률 vs 진짜이익률")
    print(f"  {target}\n")
    print(f"    매출액                    {won(engine.revenue):>18s}")
    print(f"    (−) 변동비                {won(engine.variable_cost):>18s}")
    print(f"    {'-' * 44}")
    print(f"    [1단] 공헌이익            {won(engine.contribution_margin):>18s}"
          f"   {pct(engine.contribution_margin_rate)}   ← 대표가 지금 보는 숫자")
    print(f"    (−) 배부 고정비           {won(engine.allocated_fixed):>18s}")
    print(f"    (−) 직접귀속 고정비       {won(engine.direct_fixed):>18s}")
    print(f"    (−) 맨데이 인건비         {won(engine.manday_cost):>18s}   ← ERP에 없는 원가")
    print(f"    {'-' * 44}")
    print(f"    [2단] 진짜 영업이익       {won(engine.operating_profit):>18s}"
          f"   {pct(engine.operating_profit_rate)}   ← 실제로 남는 숫자")
    drop = engine.contribution_margin_rate - engine.operating_profit_rate
    print(f"\n    이익률 {pct(engine.contribution_margin_rate)} → {pct(engine.operating_profit_rate)}"
          f" ({drop * 100:.1f}%p 하락), 사라진 금액 {won(engine.gap)}원")

    print("\n  전 현장 비교 (하락폭 순):")
    print(f"    {'프로젝트':<32s}{'공헌이익률':>10s}{'진짜이익률':>11s}{'하락폭':>9s}")
    for p in sorted(engine_rows, key=lambda p: p.operating_profit_rate - p.contribution_margin_rate):
        d = p.contribution_margin_rate - p.operating_profit_rate
        mark = "  ◀" if p.project_name == target else ""
        print(f"    {p.project_name[:30]:<32s}{pct(p.contribution_margin_rate):>10s}"
              f"{pct(p.operating_profit_rate):>11s}{d * 100:>8.1f}p{mark}")

    # ---------------------------------------------------------
    out = REPORT_DIR / f"검산_{target.replace(' ', '')}.xlsx"
    write_workbook(raw, target, hand["개월수"], out)
    rule("검산용 엑셀 생성")
    print(f"  {out.relative_to(PROJECT_ROOT)}")
    print("  시트: 검산 / 매출 / 매입 / 경비 / 맨데이 / 고정비")
    print("  '검산' 시트의 B열은 전부 SUMIF·SUMIFS·SUMPRODUCT 수식이고,")
    print("  C열이 엔진 값, D열이 그 차이다. 원본 행을 고치면 B열이 따라 움직인다.")

    conn.close()
    print()
    if not all_ok:
        sys.exit("✗ 수기 계산과 엔진 값이 어긋난다. 위 표의 ✗ 항목을 확인할 것.")
    print("✓ 전 항목 일치 — 엔진 결과는 원본 엑셀 수기 집계와 같다.")


if __name__ == "__main__":
    main()
